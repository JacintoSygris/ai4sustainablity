from typing import List

from .processor import Processor
from openpyxl import load_workbook

class ExcelProcessor(Processor):
    def extension(self) -> List[str]:
        return [".xlsx", ".xls"]

    def extract(self, path: str) -> str:
        print(f"[Excel] Processing file {path}")
        workbook = load_workbook(path)
        data_string = ""

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data_string += f"Sheet: {sheet_name}\n"  # Add sheet name to the string

            # Iterate through rows in the sheet
            for row in sheet.iter_rows(values_only=True):
                # Convert the row to a string and append it
                row_string = "\t".join([str(cell) if cell is not None else "" for cell in row])
                data_string += row_string + "\n"

            data_string += "\n"  # Add a blank line between sheets
        #print(f"[Excel] read: \n {data_string}")
        return data_string
